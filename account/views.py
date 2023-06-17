
import datetime
from email import message

import random
from .serializers import *
from .models import *
import time
from django.forms import ValidationError
from FP import settings
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from account.functionalities import limit_off, random_number
from rest_framework import viewsets
from rest_framework.views import APIView
from django.http import HttpResponse
import openpyxl
import pandas as pd
from .models import MyModel,Invoice
from django.core.mail import send_mail
from rest_framework import generics

from django_filters.rest_framework import DjangoFilterBackend
from fpdf import FPDF

from django.core.paginator import Paginator
from django.db.models import Sum,Count


class PayrollElementView(generics.ListAPIView):
    permission_classes = (IsAuthenticated,)

    queryset = Payroll.objects.all()
    serializer_class = PayrollSerializer
    filter_backends=[DjangoFilterBackend,]
    filterset_fields=['firstname','lastname', 'fathername','mothername',
                'adhar_no','pan_no']


class RegisterAPI(APIView):
    """Api to store the new user details into database"""

    def post(self, request):
        """so save the details and generating the user id"""
        serializer = UserTableSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'successfully registered'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginAPI(TokenObtainPairView):
    """Api for user to login into project"""
    permission_classes = (AllowAny,)
    serializer_class = AuthTokenSerializer
    
    
class SentMailView(APIView):
    """Api to sent the otp to user mail  id to reset the password"""

    def post(self, request):
        """sending the otp to user mail id"""
        try:
            mail = UserTable.objects.get(email=request.data['email'].lower())
        except:
            return Response({'error': 'Email does not exits.'}, status=status.HTTP_404_NOT_FOUND )
                

        if Otp.objects.filter(email=mail).exists:
            Otp.objects.filter(email=mail).delete()
        otp = Otp.objects.create(email=mail)
        otp.otp = random.randint(100000, 999999)
        otp.save()
        subject = 'Reset Your Password'
        body = f'This is your OTP to reset password {otp.otp}'


        try:
            send_mail(subject, body, settings.EMAIL_HOST_USER, [mail.email], fail_silently=False)
            return Response({"status": "mail sent "}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"status": f"An error ocured.{e} Try again!!!"}, status=status.HTTP_400_BAD_REQUEST)


class OtpVerification(APIView):
    serializer_class = OtpVerificationSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data, instance=request.data)
        if serializer.is_valid(raise_exception=True):
            return Response({"otp": "verified"}, status=status.HTTP_200_OK)
        return Response({"otp": "please generate otp again"}, status=status.HTTP_400_BAD_REQUEST)
    
class ResetPasswordview(generics.UpdateAPIView):
    """Api to reset the password and storing the new password into database"""
    if OtpVerificationSerializer:
        serializer_class = SetNewPasswordSerializer

        def post(self, request, *args, **kwargs):
            """saving the new password of the user into database"""
            serializer = self.serializer_class(data=request.data)
            serializer.is_valid(raise_exception=True)
            request_email = request.data['email']
            user_object = UserTable.objects.get(email=request_email)
            if Otp.objects.filter(email_id=user_object.pk).exists():
                if UserTable.objects.get(email=request_email):
                    user_object.password = make_password(request.data['password'])                    
                    user_object.save()
                    otp_del = Otp.objects.filter(email=user_object.id)
                    otp_del.delete()
                    return Response({'status': 'password successfully changed'}, status=status.HTTP_201_CREATED)
                return Response({'status': 'An error occured'}, status=status.HTTP_400_BAD_REQUEST)
            return Response({'status': 'An error occured'}, status=status.HTTP_400_BAD_REQUEST)

class AddAccountApi(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        query_params = request.query_params
        id = query_params['id'] if query_params.get('id') else False
        limit = query_params['limit'] if query_params.get('limit') else False
        offset  = query_params['offset'] if query_params.get('offset')  else False

        if id:
            query = Add_account.objects.filter(id=id)
        elif limit and offset:
            query = Add_account.objects.all()[int(offset):int(limit)+int(offset)]
        elif limit or offset:
            if limit:
                query = Add_account.objects.all()[:int(limit)]
            else:
                query = Add_account.objects.all()[int(offset):]        
        else:
            query = Add_account.objects.all() 

        serializer = AddAccountSerializer(query, many=True)
        return Response({'message': serializer.data}, status=status.HTTP_200_OK)       

    def post(self, request):
        serializer = AddAccountSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
       
    def post(self, request):
        serializer = AddAccountSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class BillApiView(APIView):
    permission_classes = (IsAuthenticated,)
    
    def post(self, request):
        serializer = BillSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"message": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    def get(self, request):        
        data=limit_off(Bill, request,BillSerializer)           
        
        return Response({'message': data}, status=status.HTTP_200_OK)


class InvoiceApiView(APIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):        
        data=limit_off(Invoice, request,InvoiceSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    
    def post(self, request):
        data = {
            "invoice_date":request.data['invoice_date'],
            "invoice_amount": request.data['invoice_amount'],
            "deduction": request.data['deduction'],
            "deduction_reason": request.data['deduction_reason'],
            "received_transfer":request.data['received_transfer']
        }
        var = request.data['received_transfer']
        if var == 'in':
            data["invoice_no"] = request.data['invoice_no']
            data["invoice_ref_no"] = random_number(Invoice,var,field='invoice_no')
        if var == 'out':
            data["invoice_no"] = random_number(Invoice,var,field='invoice_no')
        serializer = InvoiceSerializer(data=data)      
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            in_amount,out_amount,flag = 0,0,False
            for inv in Invoice.objects.all() :
                if inv.received_transfer == 'out':
                    out_amount += inv.invoice_amount - inv.deduction
                    flag = True
                else:
                    in_amount += inv.invoice_amount - inv.deduction
            if flag == True:
                for bill in Bill.objects.all():
                    out_amount+= bill.bill_amount
   
            return Response({"message": serializer.data, "in_amount":in_amount, "out_amount":out_amount}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PaymentApiView(APIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):       
        data=limit_off(Payment, request,PaymentSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        payment_ref_no = random_number(Payment,field='payment_ref_no',var='out')
        data = {
            "payment_date":request.data['payment_date'],
            "payment_ref_no":payment_ref_no,
            "received_payment_transfer": request.data['received_payment_transfer'],
            
        }
        serializer = PaymentSerializer(data = data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"message": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class VendorApiView(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):        
        data=limit_off(Vendor, request,VendorSerializers)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = VendorSerializers(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FinanceOutAPI(APIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):        
        data=limit_off(Finance_out, request,FinanceOutSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        ref_no = random_number(Finance_out,field='ref_no',var='out')
        data = {
            "amount":request.data['amount'],
            "invoice_detail":request.data['invoice_detail'],
            "payment_detail":request.data['payment_detail'],
            "tds_tax":request.data['tds_tax'],
            "vendor":request.data['vendor'],

            "ref_no":ref_no,
                        
        }
        serializer = FinanceOutSerializer(data=data)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response({"data":serializer.data} ,status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class Month_Finance_outApi(APIView):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):        
        data=limit_off(Month_Finance_out, request,Month_Finance_outSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        data = {
            "amount":request.data['amount'],
            "bill": request.data['bill'],
            "salary_process": request.data['salary_process'],
            
            }
                
        data["d"] = str(datetime.datetime.now()).split(" ")[0]        
        time.sleep(20)
        serializer = Month_Finance_outSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        
class FinanceInApi(APIView):
    permission_classes = (IsAuthenticated,)
    

    def get(self, request):        
        data=limit_off(Finance_in, request,FinanceInSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)



    def post(self, request):
        serializer = FinanceInSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class Financeintotal(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self, request):
        queryset= Finance_in.objects.prefetch_related().aggregate(count=Count('id'), total_amount=Sum('amount'))
        queryset2= Finance_out.objects.prefetch_related().aggregate(count=Count('id'), total_amount=Sum('amount'))

        return Response({"in":queryset,"out":queryset2})
    
class Financeouttotal(APIView):
    permission_classes = (IsAuthenticated,)
    def get(self, request):
        queryset = Finance_out.objects.aggregate(count = Count('id'),total_amount=Sum('amount'))
        return Response(queryset)


class Graduation_detailsViewSet(viewsets.ModelViewSet):
    permission_classes = (IsAuthenticated,)
    
    permission_classes = [IsAuthenticated, ]
    queryset = Graduation_details.objects.all()
    serializer_class = Graduation_detailsSerializer


class PostGraduationApi(viewsets.ModelViewSet):
    permission_classes = (IsAuthenticated,)
    
    def get(self, request):
        queryset= Finance_in.objects.prefetch_related().aggregate(count=Count('id'), total_amount=Sum('amount'))
        queryset2= Finance_out.objects.prefetch_related().aggregate(count=Count('id'), total_amount=Sum('amount'))
        return Response({'finance_in':queryset,'finance_out':queryset2})
   

class MarksheetApi(viewsets.ModelViewSet):
    permission_classes = (IsAuthenticated,)
    
    queryset = Marksheet.objects.all()
    serializer_class = MarksheetSerializer


class EvaluationAPI(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):        
        data=limit_off(Evaluation, request,EvaluationSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = EvaluationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PayrollAPI(APIView):
    permission_classes = (IsAuthenticated,)

    def get(self, request):        
        data=limit_off(Payroll, request,PayrollSerializer)                    
        return Response({'message': data}, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = PayrollSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ExcelExportView(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="finance-in.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Write headers
        headers = [
    'Amount',
    'Ref_no',
    'Invoice_no',
    'invoice_date',
    'invoice_amount',
    'deduction',
    'deduction_reason',
    'received_transfer',
    'payment_date',
    'payment_ref_no',
    'received_transfer',
    'Tds'
]
        for index, header in enumerate(headers):
            cell = chr(ord('A') + index) + '1'
            worksheet[cell] = header
        # all_fields = Finance_in._meta.fields
        # Get data
        queryset = Finance_in.objects.all()
        serializer = FinanceInSerializer(queryset, many=True)
        # Write data to Excel file
        column_mapping = [
    (1, 'amount'),
    (2, 'ref_no'),
    (3, 'invoice_detail.invoice_no'),
    (4, 'invoice_detail.invoice_date'),
    (5, 'invoice_detail.invoice_amount'),
    (6, 'invoice_detail.deduction'),
    (7, 'invoice_detail.deduction_reason'),
    (8, 'invoice_detail.received_transfer'),
    (9, 'payment_detail.payment_date'),
    (10, 'payment_detail.payment_ref_no'),
    (11, 'payment_detail.received_transfer'),
    (12, 'tds_tax')
]
        for i, row in enumerate(serializer.data, start=2):
            for column, key in column_mapping:
                value = row
                for k in key.split('.'):
                    value = value.get(k)
                    if value is None:
                        break
                worksheet.cell(row=i, column=column, value=value)
        workbook.save(response)
        return response
class ExcelExport(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="finance-out.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Write headers
        headers = [
    'Amount',
    'Ref_no',
    'Invoice_no',
    'invoice_date',
    'invoice_amount',
    'deduction',
    'deduction_reason',
    'received_transfer',
    'payment_date',
    'payment_ref_no',
    'received_transfer',
    'Tds',
    'Bill Number',
    'Bill Date',
    'Bill Amount',
    'Bill Type',
    'Salary process'
]
        for index, header in enumerate(headers):
            cell = chr(ord('A') + index) + '1'
            worksheet[cell] = header
        # Get data
        queryset = Finance_out.objects.all()
        serializer = FinanceOutSerializer(queryset, many=True)

        # Write data to Excel file
        column_mapping = [
    (1, 'amount'),
    (2, 'ref_no'),
    (3, 'invoice_detail.invoice_no'),
    (4, 'invoice_detail.invoice_date'),
    (5, 'invoice_detail.invoice_amount'),
    (6, 'invoice_detail.deduction'),
    (7, 'invoice_detail.deduction_reason'),
    (8, 'invoice_detail.received_transfer'),
    (9, 'payment_detail.payment_date'),
    (10, 'payment_detail.payment_ref_no'),
    (11, 'payment_detail.received_transfer'),
    (12, 'tds_tax'),
    (13, 'bills.bill_no'),
    (14, 'bills.bill_date'),
    (15, 'bills.bill_amount'),
    (16, 'bills.bill_type'),
    (17, 'salary_process')
]

        for i, row in enumerate(serializer.data, start=2):
            for column, key in column_mapping:
                value = row
                for k in key.split('.'):
                    value = value.get(k)
                    if value is None:
                        break
                worksheet.cell(row=i, column=column, value=value)
        workbook.save(response)
        return response

class ExcelUploadView(APIView):
    def post(self, request, format=None):
        serializer = ExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']
        df = pd.read_excel(file)  # Read the Excel file using pandas
        # Iterate over the rows and save each row as a new instance of the model
        for index, row in df.iterrows():
            MyModel.objects.create(
                column1=row['Column1'],
                column2=row['Column2'],
                # Add more fields as needed
            )
        return Response({'message': 'Data uploaded successfully'})


class FileUploadView(APIView):
    def post(self, request):
        serializer = FileUploadSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            try:
                df = pd.read_csv(file)
                print(df)
                import psycopg2
                conn = psycopg2.connect(
                    host='localhost',
                    database='payroll',
                    user='postgres',
                    password='Sourabh12'
                )
                
                from datetime import datetime
                for _, row in df.iterrows():
                    # Create AccountVendor object
                    vendor = Vendor(
                        vendor_name=row['vendor_name'],
                        vendor_address=row['vendor_address'],
                        vendor_mobileno=row['vendor_mobileno'],
                        vendor_GSTno=row['vendor_GSTno'],
                        vendor_PanCard=row['vendor_PanCard'],
                        vendor_TDS=row['vendor_TDS']
                    )
                    vendor.save()
                    # Create AccountAddAccount object
                    account = Add_account(
                        acc_no=row['acc_no'],
                        ifsc=row['ifsc'],
                        current_bal=row['current_bal'],
                        current_due=row['current_due']
                    )
                    account.save()
                    # Convert payment_date format to "YYYY-MM-DD"
                    payment_date_str = row['payment_date']
                    try:
                        payment_date_obj = datetime.strptime(payment_date_str, "%d-%m-%Y").date()
                        formatted_payment_date = payment_date_obj.strftime("%Y-%m-%d")
                        # Create AccountPayment object with formatted payment_date
                        payment = Payment(
                            payment_date=formatted_payment_date,
                            payment_ref_no=row['payment_ref_no'],
                            received_payment_transfer=row['received_payment_transfer']
                        )
                        payment.save()
                    except ValueError:
                        raise ValidationError('Invalid payment date format')
                    # Create AccountInvoice object
                    invoice_date_str = row['invoice_date']
                    try:
                        invoice_date_obj = datetime.strptime(invoice_date_str, "%d-%m-%Y").date()
                        formatted_invoice_date = invoice_date_obj.strftime("%Y-%m-%d")
                        invoice = Invoice(
                            invoice_no=row['Invoice_no'],
                            invoice_date=formatted_invoice_date,
                            invoice_amount=row['invoice_amount'],
                            deduction=row['deduction'],
                            deduction_reason=row['deduction_reason'],
                            received_transfer=row['received_transfer']
                        )
                        invoice.save()
                    except ValueError:
                        raise ValidationError('Invalid payment date format')
                    # Create FinanceIn object with foreign key relationships
                    finance_in = Finance_in(
                        amount=row['Amount'],
                        ref_no=row['Ref_no'],
                        tds_tax=row['Tds'],
                        vendor=vendor,
                        account=account,
                        invoice_detail=invoice,
                        payment_detail=payment
                    )
                    finance_in.save()
                return Response('CSV file uploaded and printed on terminal')
            except pd.errors.ParserError:
                return Response('Invalid CSV file', status=400)
        else:
            return Response(serializer.errors, status=400)



class PDF(FPDF):
    def __init__(self):
        super().__init__()
        self.WIDTH = 210
        self.HEIGHT = 297
        
    def header(self):
        self.image('assets/logo.png', 10, 8, 33)
        self.set_font('Arial', 'B', 11)
        self.cell(self.WIDTH - 80)
        self.cell(60, 1, 'Sales report', 0, 0, 'R')
        self.ln(20)
        
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128)
        self.cell(0, 10, 'Page ' + str(self.page_no()), 0, 0, 'C')

    def page_body(self, images):

        if len(images) == 3:
            self.image(images[0], 15, 25, self.WIDTH - 30)
            self.image(images[1], 15, self.WIDTH / 2 + 5, self.WIDTH - 30)
            self.image(images[2], 15, self.WIDTH / 2 + 90, self.WIDTH - 30)
        elif len(images) == 2:
            self.image(images[0], 15, 25, self.WIDTH - 30)
            self.image(images[1], 15, self.WIDTH / 2 + 5, self.WIDTH - 30)
        else:
            self.image(images[0], 15, 25, self.WIDTH - 30)
            
    def print_page(self, images):
        self.add_page()
        self.page_body(images)



class ExcelExportViewtrail(APIView):
    def get(self, request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="finance-in.xlsx"'
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Write headers
        headers = [
                    'Amount',
                    'Ref_no',
                    'Invoice_no',
                    'invoice_date',
                    'invoice_amount',
                    'deduction',
                    'deduction_reason',
                    'received_transfer',
                    'payment_date',
                    'payment_ref_no',
                    'received_transfer',
                    'Tds'
                ]
        for index, header in enumerate(headers):
            cell = chr(ord('A') + index) + '1'
            worksheet[cell] = header
        # all_fields = Finance_in._meta.fields
        # Get data
        queryset = Finance_in.objects.all()
        serializer = FinanceInSerializer(queryset, many=True)
        # Write data to Excel file
        column_mapping = [
                    (1, 'amount'),
                    (2, 'ref_no'),
                    (3, 'invoice_detail.invoice_no'),
                    (4, 'invoice_detail.invoice_date'),
                    (5, 'invoice_detail.invoice_amount'),
                    (6, 'invoice_detail.deduction'),
                    (7, 'invoice_detail.deduction_reason'),
                    (8, 'invoice_detail.received_transfer'),
                    (9, 'payment_detail.payment_date'),
                    (10, 'payment_detail.payment_ref_no'),
                    (11, 'payment_detail.received_transfer'),
                    (12, 'tds_tax')
]
        for i, row in enumerate(serializer.data, start=2):
            for column, key in column_mapping:
                value = row
                for k in key.split('.'):
                    value = value.get(k)
                    if value is None:
                        break
                worksheet.cell(row=i, column=column, value=value)
        workbook.save(response)
        return response

